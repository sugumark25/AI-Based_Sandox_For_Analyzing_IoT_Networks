# backend/realtime_collector.py
"""
realtime_collector.py
=====================
Stores ESP32 flow data into SEPARATE files:

    dataset/esp32_attacks.csv      -- attack flows only
    dataset/esp32_normal.csv       -- normal flows (deduplicated by pattern)
    dataset/esp32_sensors.csv      -- DHT22 sensor readings
    dataset/realtime_esp32.csv     -- ALL flows combined (legacy)
    dataset/realtime_esp32.xlsx    -- Excel with 4 sheets (Raw, Stats, Attacks, Normal)

Normal flow deduplication:
    Uses packet_rate bucket + byte_rate bucket + proto + conn_state as key.
    Only saves a normal flow if its pattern has NOT been seen before.
    This prevents thousands of identical idle flows filling the CSV.

MQTT Topics:
    iot/edge/attacks/<device_id>   attack flows
    iot/edge/normal/<device_id>    normal flows
    iot/sensors/<device_id>        DHT22 sensor readings
"""

import os
import sys
import csv
import json
import time
import signal
import logging
import argparse
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("realtime_collector")

# ── Paths ─────────────────────────────────────────────────────────
BASE_DIR          = Path(__file__).parent
DATASET_DIR       = BASE_DIR / "dataset"
REALTIME_CSV_PATH = DATASET_DIR / "realtime_esp32.csv"
REALTIME_XLS_PATH = DATASET_DIR / "realtime_esp32.xlsx"
ATTACK_CSV_PATH   = DATASET_DIR / "esp32_attacks.csv"
NORMAL_CSV_PATH   = DATASET_DIR / "esp32_normal.csv"
SENSOR_CSV_PATH   = DATASET_DIR / "esp32_sensors.csv"
SENSOR_XLS_PATH   = DATASET_DIR / "esp32_sensors.xlsx"

# ── MQTT ──────────────────────────────────────────────────────────
MQTT_BROKER        = "localhost"
MQTT_PORT          = 1883
MQTT_TOPIC_ATTACKS = "iot/edge/attacks/#"
MQTT_TOPIC_NORMAL  = "iot/edge/normal/#"
MQTT_TOPIC_SENSORS = "iot/sensors/#"

# ── Sensor thresholds ─────────────────────────────────────────────
TEMP_ALERT_C  = 38.0
HUM_ALERT_PCT = 80.0

MAX_ROWS_BEFORE_ROTATE = 50_000

# ── Excel sheet names ─────────────────────────────────────────────
SHEET_RAW     = "RawData"
SHEET_STATS   = "Stats"
SHEET_ATTACKS = "Attacks"
SHEET_NORMAL  = "NormalFlows"

SENSOR_SHEET_DATA   = "SensorData"
SENSOR_SHEET_STATS  = "Stats"
SENSOR_SHEET_ALERTS = "Alerts"

COLUMNS = [
    ("Timestamp",         16, "meta"),
    ("Device ID",         11, "meta"),
    ("Edge Decision",     14, "label"),
    ("Confidence",        11, "label"),
    ("Z-Score",           10, "label"),
    ("Inference (ms)",    14, "meta"),
    ("Duration (s)",      12, "feat"),
    ("Src Bytes",         11, "feat"),
    ("Dst Bytes",         11, "feat"),
    ("Src Pkts",          10, "feat"),
    ("Dst Pkts",          10, "feat"),
    ("Packet Rate",       12, "feat"),
    ("Byte Rate",         12, "feat"),
    ("Bytes/Pkt",         11, "feat"),
    ("Payload Ratio",     12, "feat"),
    ("Proto TCP",         10, "proto"),
    ("Proto UDP",         10, "proto"),
    ("Proto ICMP",        10, "proto"),
    ("Conn OK",            9, "conn"),
    ("Conn S0",            9, "conn"),
    ("Conn REJ",           9, "conn"),
    ("Jitter",            10, "feat"),
    ("Flow Weight",       11, "feat"),
    ("Magnitude",         11, "feat"),
    ("Variance",          10, "feat"),
    ("Logged In",         10, "feat"),
    ("Failed Logins",     13, "feat"),
    ("Srv Count",         10, "feat"),
]

SENSOR_COLUMNS = [
    ("Timestamp",       18),
    ("Device ID",       14),
    ("Temp (C)",        11),
    ("Humidity (%)",    13),
    ("Heat Index (C)",  16),
    ("Label",           10),
]

HDR_COLORS = {
    "meta":  "1F3864",
    "label": "6C3483",
    "feat":  "154360",
    "proto": "145A32",
    "conn":  "6E2F1A",
}

ROW_CLR = {
    "attack":  "FADBD8",
    "normal":  "D5F5E3",
}

SENSOR_ROW_CLR = {
    "HOT":    "FADBD8",
    "WARM":   "FDEBD0",
    "HUMID":  "F9E79F",
    "COOL":   "D6EAF8",
    "NORMAL": "D5F5E3",
}

FEATURE_NAMES = [
    "duration", "src_bytes", "dst_bytes", "src_pkts", "dst_pkts",
    "packet_rate", "byte_rate", "bytes_per_pkt", "payload_ratio",
    "proto_tcp", "proto_udp", "proto_icmp",
    "conn_ok", "conn_s0", "conn_rej",
    "jitter", "flow_weight", "magnitude", "variance",
    "logged_in", "num_failed_logins", "srv_count",
]

CSV_COLUMNS = (
    ["timestamp", "device_id", "edge_decision",
     "edge_confidence", "z_score", "inference_ms"]
    + FEATURE_NAMES
)

SENSOR_CSV_COLUMNS = ["timestamp", "device_id", "temp", "hum", "heatIndex", "label"]

# ── State ─────────────────────────────────────────────────────────
_lock                = threading.Lock()
_queue               = []
_sensor_queue        = []
_records_written     = 0
_attack_count        = 0
_normal_count        = 0
_normal_dedup_count  = 0
_device_counts       = {}
_sensor_written      = 0
_sensor_alert_count  = 0
_sensor_devices      = {}
_running             = True
_FLUSH_ROWS          = 50
_FLUSH_SECS          = 10

EXCEL_ENABLED = True

# ── Normal flow deduplication ─────────────────────────────────────
_seen_normal_patterns: set = set()

# ── ESP32 connection tracking ─────────────────────────────────────
_last_esp32_message   = 0.0
_esp32_connected      = False
_esp32_ever_seen      = False          # True once the first message arrives
ESP32_TIMEOUT_SECS    = 30
ESP32_GRACE_SECS      = 60            # Grace period before "NEVER CONNECTED" warning


def _normal_pattern_key(features_dict: dict) -> tuple:
    def bucket(v, thresholds):
        for i, t in enumerate(thresholds):
            if v < t:
                return i
        return len(thresholds)

    pkt_rate  = float(features_dict.get("packet_rate", 0))
    byte_rate = float(features_dict.get("byte_rate",   0))
    pkt_b     = bucket(pkt_rate,  [10, 50, 100, 250, 500, 1000, 2000])
    byte_b    = bucket(byte_rate, [1000, 10000, 50000, 100000, 500000])
    proto = (
        int(features_dict.get("proto_tcp",  0)),
        int(features_dict.get("proto_udp",  0)),
        int(features_dict.get("proto_icmp", 0)),
    )
    conn = (
        int(features_dict.get("conn_ok",  0)),
        int(features_dict.get("conn_s0",  0)),
        int(features_dict.get("conn_rej", 0)),
    )
    return (pkt_b, byte_b, proto, conn)


# ─────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────

def _bd():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_header(ws, cols):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    bd = _bd()
    for i, (name, width, grp) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", start_color=HDR_COLORS.get(grp, "1F3864"))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bd
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _init_stats_sheet(ws):
    bd     = _bd()
    hfill  = PatternFill("solid", start_color="1F3864")
    vfill  = PatternFill("solid", start_color="EBF5FB")
    tf     = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    lf     = Font(name="Arial", bold=True, size=10)
    vf     = Font(name="Arial", size=10, color="154360")
    center = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    def title(r, txt):
        c = ws.cell(row=r, column=1, value=txt)
        c.font = tf; c.fill = hfill
        c.alignment = Alignment(horizontal="left"); c.border = bd
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 20

    def stat(r, label, val="--"):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        lc.font = lf; lc.border = bd
        vc.font = vf; vc.fill = vfill; vc.border = bd; vc.alignment = center

    title(1,  "ESP32 Real-Time Collector -- Live Stats")
    title(3,  "Session Summary")
    stat(4,   "Started At",              datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    stat(5,   "Last Updated",            "--")
    stat(6,   "Total Records",           0)
    stat(7,   "Attack Records",          0)
    stat(8,   "Normal Records (saved)",  0)
    stat(9,   "Normal Deduplicated",     0)
    stat(10,  "Unique Normal Patterns",  0)
    stat(11,  "Attack Rate",             "0.00%")
    stat(12,  "ESP32 Status",            "WAITING")
    title(14, "Per-Device Counts")
    for col, label in [(1, "Device ID"), (2, "Flow Count")]:
        c = ws.cell(row=15, column=col, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = hfill; c.border = bd; c.alignment = center


def _update_stats_sheet(wb):
    ws     = wb[SHEET_STATS]
    bd     = _bd()
    center = Alignment(horizontal="center")
    vf     = Font(name="Arial", size=10, color="154360")
    vfill  = PatternFill("solid", start_color="EBF5FB")
    lf     = Font(name="Arial", bold=True, size=9)

    esp32_status = "CONNECTED" if _esp32_connected else (
        "DISCONNECTED" if _esp32_ever_seen else "NEVER CONNECTED"
    )

    def setval(r, v):
        c = ws.cell(row=r, column=2, value=v)
        c.font = vf; c.fill = vfill; c.border = bd; c.alignment = center

    setval(5,  datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    setval(6,  _records_written)
    setval(7,  _attack_count)
    setval(8,  _normal_count)
    setval(9,  _normal_dedup_count)
    setval(10, len(_seen_normal_patterns))
    setval(11, f"{(_attack_count / max(_records_written, 1)) * 100:.2f}%")
    setval(12, esp32_status)

    for idx, (dev, cnt) in enumerate(_device_counts.items()):
        r  = 16 + idx
        lc = ws.cell(row=r, column=1, value=dev)
        vc = ws.cell(row=r, column=2, value=cnt)
        lc.font = lf; lc.border = bd
        vc.font = vf; vc.fill = vfill; vc.border = bd; vc.alignment = center


def _create_excel():
    DATASET_DIR.mkdir(parents=True, exist_ok=True)
    wb  = Workbook()
    ws1 = wb.active; ws1.title = SHEET_RAW
    _write_header(ws1, COLUMNS)
    ws2 = wb.create_sheet(SHEET_STATS)
    _init_stats_sheet(ws2)
    ws3 = wb.create_sheet(SHEET_ATTACKS)
    _write_header(ws3, COLUMNS)
    ws4 = wb.create_sheet(SHEET_NORMAL)
    _write_header(ws4, COLUMNS)
    wb.save(str(REALTIME_XLS_PATH))
    log.info("Created flow Excel: %s", REALTIME_XLS_PATH)


def _write_sensor_header(ws):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    bd    = _bd()
    hfill = PatternFill("solid", start_color="1F3864")
    for i, (name, width) in enumerate(SENSOR_COLUMNS, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill      = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bd
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SENSOR_COLUMNS))}1"


def _init_sensor_stats_sheet(ws):
    bd    = _bd()
    hfill = PatternFill("solid", start_color="1F3864")
    vfill = PatternFill("solid", start_color="EBF5FB")
    tf    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    lf    = Font(name="Arial", bold=True, size=10)
    vf    = Font(name="Arial", size=10, color="154360")
    ctr   = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    def title(r, txt):
        c = ws.cell(row=r, column=1, value=txt)
        c.font = tf; c.fill = hfill
        c.alignment = Alignment(horizontal="left"); c.border = bd
        ws.merge_cells(f"A{r}:B{r}")
        ws.row_dimensions[r].height = 20

    def stat(r, label, val="--"):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=val)
        lc.font = lf; lc.border = bd
        vc.font = vf; vc.fill = vfill; vc.border = bd; vc.alignment = ctr

    title(1, "ESP32 Sensor Collector -- Live Stats")
    title(3, "Session Summary")
    stat(4, "Started At",     datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    stat(5, "Last Updated",   "--")
    stat(6, "Total Readings", 0)
    stat(7, "Alert Readings", 0)
    stat(8, "Alert Rate",     "0.00%")
    stat(9, "ESP32 Status",   "WAITING")
    title(11, "Per-Device Counts")
    for col, label in [(1, "Device ID"), (2, "Reading Count")]:
        c = ws.cell(row=12, column=col, value=label)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = hfill; c.border = bd
        c.alignment = Alignment(horizontal="center")


def _update_sensor_stats_sheet(wb):
    ws    = wb[SENSOR_SHEET_STATS]
    bd    = _bd()
    ctr   = Alignment(horizontal="center")
    vf    = Font(name="Arial", size=10, color="154360")
    vfill = PatternFill("solid", start_color="EBF5FB")
    lf    = Font(name="Arial", bold=True, size=9)

    esp32_status = "CONNECTED" if _esp32_connected else (
        "DISCONNECTED" if _esp32_ever_seen else "NEVER CONNECTED"
    )

    def setval(r, v):
        c = ws.cell(row=r, column=2, value=v)
        c.font = vf; c.fill = vfill; c.border = bd; c.alignment = ctr

    setval(5, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    setval(6, _sensor_written)
    setval(7, _sensor_alert_count)
    setval(8, f"{(_sensor_alert_count / max(_sensor_written, 1)) * 100:.2f}%")
    setval(9, esp32_status)

    for idx, (dev, cnt) in enumerate(_sensor_devices.items()):
        r  = 13 + idx
        lc = ws.cell(row=r, column=1, value=dev)
        vc = ws.cell(row=r, column=2, value=cnt)
        lc.font = lf; lc.border = bd
        vc.font = vf; vc.fill = vfill; vc.border = bd; vc.alignment = ctr


def _create_sensor_excel():
    DATASET_DIR.mkdir(parents=True, exist_ok=True)
    wb  = Workbook()
    ws1 = wb.active; ws1.title = SENSOR_SHEET_DATA
    _write_sensor_header(ws1)
    ws2 = wb.create_sheet(SENSOR_SHEET_STATS)
    _init_sensor_stats_sheet(ws2)
    ws3 = wb.create_sheet(SENSOR_SHEET_ALERTS)
    _write_sensor_header(ws3)
    wb.save(str(SENSOR_XLS_PATH))
    log.info("Created sensor Excel: %s", SENSOR_XLS_PATH)


# ─────────────────────────────────────────────────────────────────
# CSV helpers
# ─────────────────────────────────────────────────────────────────

def _ensure_dataset_dir():
    DATASET_DIR.mkdir(parents=True, exist_ok=True)


def _ensure_csv_headers():
    _ensure_dataset_dir()
    for path in [REALTIME_CSV_PATH, ATTACK_CSV_PATH, NORMAL_CSV_PATH]:
        if not path.exists():
            with open(str(path), "w", newline="", encoding="utf-8") as f:
                csv.DictWriter(f, fieldnames=CSV_COLUMNS).writeheader()
            log.info("Created CSV: %s", path.name)


def _rotate_if_needed():
    if not REALTIME_CSV_PATH.exists():
        return
    with open(str(REALTIME_CSV_PATH), "r") as f:
        row_count = sum(1 for _ in f) - 1
    if row_count >= MAX_ROWS_BEFORE_ROTATE:
        ts      = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        archive = DATASET_DIR / f"realtime_esp32_{ts}.csv"
        REALTIME_CSV_PATH.rename(archive)
        log.info("CSV rotated -> %s", archive)
        _ensure_csv_headers()


# ─────────────────────────────────────────────────────────────────
# Core flow write & flush
# ─────────────────────────────────────────────────────────────────

def _build_row(payload: dict):
    features_list = payload.get("features", [])
    if len(features_list) != 22:
        features_list = (features_list + [0.0] * 22)[:22]

    features_dict = dict(zip(FEATURE_NAMES, features_list))
    is_attack     = bool(payload.get("edge_decision", False))
    decision_str  = "attack" if is_attack else "normal"
    ts            = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    device_id     = str(payload.get("device_id",       "unknown"))
    confidence    = round(float(payload.get("edge_confidence", 0.0)), 4)
    z_score       = round(float(payload.get("z_score",         0.0)), 4)
    infer_ms      = round(float(payload.get("inference_ms",    0.0)), 2)

    csv_row = {
        "timestamp":       ts,
        "device_id":       device_id,
        "edge_decision":   decision_str,
        "edge_confidence": confidence,
        "z_score":         z_score,
        "inference_ms":    infer_ms,
        **features_dict,
    }
    excel_row = (
        [ts, device_id, decision_str, confidence, z_score, infer_ms]
        + [round(float(features_dict[f]), 6) for f in FEATURE_NAMES]
    )
    return csv_row, excel_row, is_attack, features_dict, device_id, z_score, confidence


def flush_to_excel():
    global _queue
    if not _queue:
        return
    if not EXCEL_ENABLED:
        _queue.clear()
        return
    rows = _queue[:]
    _queue.clear()

    try:
        if not REALTIME_XLS_PATH.exists():
            _create_excel()
        wb  = load_workbook(str(REALTIME_XLS_PATH))
        ws1 = wb[SHEET_RAW]
        ws3 = wb[SHEET_ATTACKS]
        ws4 = wb[SHEET_NORMAL]
        nr  = ws1.max_row + 1
        na  = ws3.max_row + 1
        nn  = ws4.max_row + 1
        bd  = _bd()

        for excel_row, decision in rows:
            fill = PatternFill("solid", start_color=ROW_CLR.get(decision, "F2F3F4"))
            for col, val in enumerate(excel_row, 1):
                c = ws1.cell(row=nr, column=col, value=val)
                c.fill = fill; c.border = bd
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(
                    horizontal="center" if col <= 3 else "right")
                if col > 3:
                    c.number_format = "0.0000"
                if col == 3:
                    clr = "922B21" if decision == "attack" else "1D6A39"
                    c.font = Font(name="Arial", bold=(decision=="attack"),
                                  size=9, color=clr)
            if decision == "attack":
                for col, val in enumerate(excel_row, 1):
                    c = ws3.cell(row=na, column=col, value=val)
                    c.fill = fill; c.border = bd
                    c.font = Font(name="Arial", size=9)
                na += 1
            else:
                for col, val in enumerate(excel_row, 1):
                    c = ws4.cell(row=nn, column=col, value=val)
                    c.fill = fill; c.border = bd
                    c.font = Font(name="Arial", size=9)
                nn += 1
            nr += 1

        _update_stats_sheet(wb)
        wb.save(str(REALTIME_XLS_PATH))
        mb = REALTIME_XLS_PATH.stat().st_size / 1_048_576
        log.info("Flow Excel saved: total=%d  %.2fMB", _records_written, mb)

    except Exception as e:
        log.error("Flow Excel write error: %s", e)
        _queue = list(rows) + _queue


def _do_flush():
    if not EXCEL_ENABLED:
        return
    with _lock:
        flush_to_excel()
        _flush_sensor_excel()


def _flush_loop():
    while _running:
        time.sleep(_FLUSH_SECS)
        threading.Thread(target=_do_flush, daemon=True).start()


def write_realtime_record(payload: dict):
    global _records_written, _attack_count, _normal_count
    global _normal_dedup_count, _seen_normal_patterns

    csv_row, excel_row, is_attack, features_dict, device_id, z_score, confidence = \
        _build_row(payload)
    decision_str = "attack" if is_attack else "normal"

    with _lock:
        if not is_attack:
            pattern_key = _normal_pattern_key(features_dict)
            if pattern_key in _seen_normal_patterns:
                _normal_dedup_count += 1
                return
            _seen_normal_patterns.add(pattern_key)

        _rotate_if_needed()

        write_hdr = not REALTIME_CSV_PATH.exists()
        with open(str(REALTIME_CSV_PATH), "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
            if write_hdr: w.writeheader()
            w.writerow(csv_row)

        target_path = ATTACK_CSV_PATH if is_attack else NORMAL_CSV_PATH
        hdr = not target_path.exists()
        with open(str(target_path), "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
            if hdr: w.writeheader()
            w.writerow(csv_row)

        _queue.append((excel_row, decision_str))
        _records_written += 1
        if is_attack:
            _attack_count += 1
        else:
            _normal_count += 1
        _device_counts[device_id] = _device_counts.get(device_id, 0) + 1
        if len(_queue) >= _FLUSH_ROWS:
            flush_to_excel()

    label_tag = "ATTACK" if is_attack else "normal"
    log.info("Flow #%05d [%-7s] device=%-10s z=%.3f conf=%.3f "
             "atk=%d norm=%d dedup=%d patterns=%d",
             _records_written, label_tag, device_id,
             z_score, confidence,
             _attack_count, _normal_count,
             _normal_dedup_count, len(_seen_normal_patterns))


# ─────────────────────────────────────────────────────────────────
# Sensor write & flush
# ─────────────────────────────────────────────────────────────────

def _get_sensor_label(temp: float, hum: float) -> str:
    if temp >= 40:          return "HOT"
    if temp >= 35:          return "WARM"
    if hum > HUM_ALERT_PCT: return "HUMID"
    if temp < 20:           return "COOL"
    return "NORMAL"


def _flush_sensor_excel():
    global _sensor_queue
    if not _sensor_queue:
        return
    if not EXCEL_ENABLED:
        _sensor_queue.clear()
        return
    rows = _sensor_queue[:]
    _sensor_queue.clear()

    try:
        if not SENSOR_XLS_PATH.exists():
            _create_sensor_excel()
        wb  = load_workbook(str(SENSOR_XLS_PATH))
        ws1 = wb[SENSOR_SHEET_DATA]
        ws3 = wb[SENSOR_SHEET_ALERTS]
        nr  = ws1.max_row + 1
        na  = ws3.max_row + 1
        bd  = _bd()

        for row_vals, label, is_alert in rows:
            fill = PatternFill("solid", start_color=SENSOR_ROW_CLR.get(label, "F2F3F4"))
            for col, val in enumerate(row_vals, 1):
                c = ws1.cell(row=nr, column=col, value=val)
                c.fill = fill; c.border = bd
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(
                    horizontal="center" if col <= 2 else "right")
                if col in (3, 4, 5):
                    c.number_format = "0.0"
            if is_alert:
                for col, val in enumerate(row_vals, 1):
                    c = ws3.cell(row=na, column=col, value=val)
                    c.fill = fill; c.border = bd
                    c.font = Font(name="Arial", size=9)
                na += 1
            nr += 1

        _update_sensor_stats_sheet(wb)
        wb.save(str(SENSOR_XLS_PATH))
        mb = SENSOR_XLS_PATH.stat().st_size / 1_048_576
        log.info("Sensor Excel saved: total=%d  %.2fMB", _sensor_written, mb)

    except Exception as e:
        log.error("Sensor Excel write error: %s", e)
        _sensor_queue = list(rows) + _sensor_queue


def write_sensor_record(payload: dict):
    global _sensor_written, _sensor_alert_count

    ts        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    device_id = str(payload.get("device_id", "ESP32"))
    temp      = round(float(payload.get("temp_c",       payload.get("temp",      0.0))), 1)
    hum       = round(float(payload.get("humidity",     payload.get("hum",       0.0))), 1)
    heat_idx  = round(float(payload.get("heat_index_c", payload.get("heatIndex", 0.0))), 1)
    label     = _get_sensor_label(temp, hum)
    is_alert  = temp >= TEMP_ALERT_C or hum >= HUM_ALERT_PCT
    row_vals  = [ts, device_id, temp, hum, heat_idx, label]

    with _lock:
        DATASET_DIR.mkdir(parents=True, exist_ok=True)
        write_hdr = not SENSOR_CSV_PATH.exists()
        with open(str(SENSOR_CSV_PATH), "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=SENSOR_CSV_COLUMNS)
            if write_hdr: w.writeheader()
            w.writerow({"timestamp": ts, "device_id": device_id,
                        "temp": temp, "hum": hum,
                        "heatIndex": heat_idx, "label": label})
        _sensor_queue.append((row_vals, label, is_alert))
        _sensor_written += 1
        if is_alert:
            _sensor_alert_count += 1
        _sensor_devices[device_id] = _sensor_devices.get(device_id, 0) + 1
        if len(_sensor_queue) >= _FLUSH_ROWS:
            _flush_sensor_excel()

    alert_tag = "ALERT" if is_alert else "OK"
    log.info("Sensor [%s] %s Temp=%.1fC Hum=%.1f%% %s total=%d",
             device_id, alert_tag, temp, hum, label, _sensor_written)


# ─────────────────────────────────────────────────────────────────
# ESP32 watchdog  ← UPDATED
# ─────────────────────────────────────────────────────────────────

def _esp32_watchdog_loop():
    """
    Monitors ESP32 heartbeat and manages _esp32_connected state.

    States:
      WAITING         — no message ever received yet (within grace period)
      NEVER CONNECTED — grace period elapsed, still no message
      CONNECTED       — messages arriving within ESP32_TIMEOUT_SECS
      DISCONNECTED    — was connected, but timed out
    """
    global _esp32_connected, _last_esp32_message, _esp32_ever_seen

    _startup_time = time.time()
    _never_connected_warned = False         # warn only once per startup

    while _running:
        time.sleep(5)
        now = time.time()

        # ── Case 1: currently marked connected — check for timeout ──
        if _esp32_connected:
            if _last_esp32_message > 0:
                elapsed = now - _last_esp32_message
                if elapsed > ESP32_TIMEOUT_SECS:
                    _esp32_connected = False
                    log.warning(
                        "ESP32 DISCONNECTED — no message for %.0fs "
                        "— data collection paused", elapsed
                    )

        # ── Case 2: currently disconnected — check for reconnect ────
        elif _esp32_ever_seen and _last_esp32_message > 0:
            elapsed = now - _last_esp32_message
            if elapsed <= ESP32_TIMEOUT_SECS:
                _esp32_connected = True
                log.info("ESP32 RECONNECTED — data collection resumed")

        # ── Case 3: never seen at all — warn after grace period ─────
        elif not _esp32_ever_seen:
            startup_elapsed = now - _startup_time
            if startup_elapsed > ESP32_GRACE_SECS and not _never_connected_warned:
                _never_connected_warned = True
                log.warning(
                    "ESP32 NEVER CONNECTED — no message received in %.0fs. "
                    "Check: device power, WiFi signal, MQTT broker (%s:%d), "
                    "topic names (iot/edge/attacks/#, iot/edge/normal/#, iot/sensors/#)",
                    startup_elapsed, MQTT_BROKER, MQTT_PORT
                )


# ─────────────────────────────────────────────────────────────────
# MQTT listener
# ─────────────────────────────────────────────────────────────────

def _on_connect(client, userdata, flags, rc):
    if rc == 0:
        client.subscribe(MQTT_TOPIC_ATTACKS, qos=1)
        client.subscribe(MQTT_TOPIC_NORMAL,  qos=1)
        client.subscribe(MQTT_TOPIC_SENSORS, qos=1)
        log.info("MQTT connected -> subscribed to attacks + normal + sensors")
    else:
        log.error("MQTT connection failed: rc=%d", rc)


def _on_message(client, userdata, msg):
    global _last_esp32_message, _esp32_connected, _esp32_ever_seen
    try:
        # ── Update ESP32 heartbeat ────────────────────────────
        _last_esp32_message = time.time()

        if not _esp32_ever_seen:
            _esp32_ever_seen = True
            log.info("ESP32 FIRST MESSAGE RECEIVED — device is online")

        if not _esp32_connected:
            _esp32_connected = True
            log.info("ESP32 CONNECTED — data collection active")

        payload = json.loads(msg.payload.decode("utf-8"))

        if "sensors" in msg.topic:
            parts = msg.topic.split("/")
            if "device_id" not in payload:
                payload["device_id"] = parts[-1] if len(parts) >= 3 else "ESP32"
            write_sensor_record(payload)
            return

        if "normal" in msg.topic:
            payload["edge_decision"] = False
        elif "attacks" in msg.topic:
            payload["edge_decision"] = True

        if "device_id" not in payload:
            parts = msg.topic.split("/")
            payload["device_id"] = parts[-1] if len(parts) >= 3 else "unknown"

        write_realtime_record(payload)

    except json.JSONDecodeError as e:
        log.warning("JSON decode error on %s: %s", msg.topic, e)
    except Exception as e:
        log.warning("Message error: %s", e)


def _on_disconnect(client, userdata, rc):
    if rc != 0:
        log.warning("MQTT disconnected (rc=%d) -- will auto-reconnect", rc)


def start_mqtt_listener(broker: str = MQTT_BROKER, port: int = MQTT_PORT):
    try:
        import paho.mqtt.client as mqtt
    except ImportError:
        log.error("paho-mqtt not installed — run: pip install paho-mqtt")
        return

    try:
        client = mqtt.Client(
            client_id="realtime_collector",
            callback_api_version=mqtt.CallbackAPIVersion.VERSION1)
    except AttributeError:
        client = mqtt.Client(client_id="realtime_collector")

    client.on_connect    = _on_connect
    client.on_message    = _on_message
    client.on_disconnect = _on_disconnect
    client.reconnect_delay_set(min_delay=2, max_delay=30)

    log.info("Connecting to MQTT %s:%d ...", broker, port)
    while _running:
        try:
            client.connect(broker, port, keepalive=60)
            client.loop_forever()
            break
        except Exception as e:
            if not _running: break
            log.warning("MQTT retry in 5s: %s", e)
            time.sleep(5)


start_mqtt = start_mqtt_listener


def start_http_listener(host: str = "0.0.0.0", port: int = 5001):
    try:
        from flask import Flask, request, jsonify
    except ImportError:
        log.error("flask not installed — run: pip install flask")
        return

    http_app = Flask("realtime_collector")

    @http_app.post("/api/realtime")
    def api_realtime():
        global _last_esp32_message, _esp32_connected, _esp32_ever_seen
        _last_esp32_message = time.time()
        if not _esp32_ever_seen:
            _esp32_ever_seen = True
            log.info("ESP32 FIRST MESSAGE via HTTP — device is online")
        if not _esp32_connected:
            _esp32_connected = True
            log.info("ESP32 CONNECTED via HTTP — data collection active")
        raw = request.get_json(force=True, silent=True) or {}
        write_realtime_record(raw)
        return jsonify({"stored": True, "total": _records_written,
                        "attacks": _attack_count, "normal": _normal_count})

    @http_app.post("/sensor")
    def api_sensor():
        global _last_esp32_message, _esp32_connected, _esp32_ever_seen
        _last_esp32_message = time.time()
        if not _esp32_ever_seen:
            _esp32_ever_seen = True
        if not _esp32_connected:
            _esp32_connected = True
        raw = request.get_json(force=True, silent=True) or {}
        write_sensor_record(raw)
        return jsonify({"stored": True, "total": _sensor_written})

    # 🔥 Alias route to fix 404 for /api/stats
    @http_app.get("/api/stats")
    def api_stats_alias():
        return jsonify(get_stats())

    @http_app.get("/api/realtime/recent")
    def api_recent():
        n = int(request.args.get("n", 50))
        return jsonify(get_recent_records(n))

    @http_app.get("/api/sensor/stats")
    def api_sensor_stats():
        return jsonify({
            "sensor_total":  _sensor_written,
            "sensor_alerts": _sensor_alert_count,
        })

    @http_app.get("/api/sensor/recent")
    def api_sensor_recent():
        n = int(request.args.get("n", 30))
        if not SENSOR_CSV_PATH.exists():
            return jsonify([])
        with open(str(SENSOR_CSV_PATH), newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        return jsonify(rows[-n:])

    log.info("HTTP listener on http://%s:%d", host, port)
    http_app.run(host=host, port=port, debug=False, use_reloader=False)


start_http = start_http_listener


def ingest_from_http(payload: dict) -> bool:
    try:
        write_realtime_record(payload)
        return True
    except Exception as e:
        log.error("ingest_from_http error: %s", e)
        return False


def get_stats() -> dict:
    mb = REALTIME_XLS_PATH.stat().st_size / 1_048_576 \
         if REALTIME_XLS_PATH.exists() else 0.0
    esp32_status = "CONNECTED" if _esp32_connected else (
        "DISCONNECTED" if _esp32_ever_seen else "NEVER CONNECTED"
    )
    return {
        "total_records":          _records_written,
        "attack_records":         _attack_count,
        "normal_records":         _normal_count,
        "normal_deduplicated":    _normal_dedup_count,
        "unique_normal_patterns": len(_seen_normal_patterns),
        "attack_rate_pct":        round(_attack_count / max(_records_written, 1) * 100, 2),
        "excel_size_mb":          round(mb, 3),
        "sensor_total":           _sensor_written,
        "sensor_alerts":          _sensor_alert_count,
        "esp32_status":           esp32_status,
        "esp32_ever_seen":        _esp32_ever_seen,
        "esp32_last_seen_secs":   round(time.time() - _last_esp32_message, 1)
                                  if _last_esp32_message > 0 else None,
    }


def get_recent_records(n: int = 50) -> list:
    if not REALTIME_CSV_PATH.exists():
        return []
    with _lock:
        with open(str(REALTIME_CSV_PATH), "r", newline="", encoding="utf-8") as f:
            records = list(csv.DictReader(f))
    return records[-n:]


def _stats_loop(interval_sec: int = 30):
    while _running:
        time.sleep(interval_sec)
        mb_flow   = REALTIME_XLS_PATH.stat().st_size / 1_048_576 \
                    if REALTIME_XLS_PATH.exists() else 0.0
        mb_sensor = SENSOR_XLS_PATH.stat().st_size / 1_048_576 \
                    if SENSOR_XLS_PATH.exists() else 0.0
        esp32_status = "CONNECTED" if _esp32_connected else (
            "DISCONNECTED" if _esp32_ever_seen else "NEVER CONNECTED"
        )
        log.info(
            "STATS [ESP32:%s] flows=%d(atk=%d norm=%d dedup=%d) "
            "sensors=%d flow=%.2fMB sensor=%.2fMB",
            esp32_status,
            _records_written, _attack_count, _normal_count,
            _normal_dedup_count,
            _sensor_written, mb_flow, mb_sensor
        )


def shutdown(sig=None, frame=None):
    global _running
    _running = False
    log.info("Shutting down...")
    with _lock:
        flush_to_excel()
        _flush_sensor_excel()
    log.info("Final: flows=%d(atk=%d norm=%d) sensors=%d",
             _records_written, _attack_count, _normal_count, _sensor_written)
    sys.exit(0)


def start_realtime_collector(broker: str = MQTT_BROKER, port: int = MQTT_PORT):
    _ensure_csv_headers()
    if not REALTIME_XLS_PATH.exists():
        _create_excel()
    if not SENSOR_XLS_PATH.exists():
        _create_sensor_excel()

    log.info("Realtime collector starting...")
    log.info("  Attack CSV   : %s", ATTACK_CSV_PATH)
    log.info("  Normal CSV   : %s (deduplicated by pattern)", NORMAL_CSV_PATH)
    log.info("  Sensor CSV   : %s", SENSOR_CSV_PATH)
    log.info("  Combined CSV : %s", REALTIME_CSV_PATH)
    log.info("  ESP32 timeout: %ds  grace: %ds", ESP32_TIMEOUT_SECS, ESP32_GRACE_SECS)

    threading.Thread(target=_flush_loop,          daemon=True).start()
    threading.Thread(target=_stats_loop,          daemon=True).start()
    threading.Thread(target=_esp32_watchdog_loop, daemon=True).start()
    start_mqtt_listener(broker, port)


def main():
    ap = argparse.ArgumentParser(description="ESP32 Real-Time Collector")
    ap.add_argument("--broker",    default=MQTT_BROKER)
    ap.add_argument("--port",      type=int, default=MQTT_PORT)
    ap.add_argument("--http-host", default="0.0.0.0")
    ap.add_argument("--http-port", type=int, default=5001)
    ap.add_argument("--no-mqtt",   action="store_true")
    ap.add_argument("--no-http",   action="store_true")
    args = ap.parse_args()

    signal.signal(signal.SIGINT,  shutdown)
    signal.signal(signal.SIGTERM, shutdown)

    print("\n" + "=" * 62)
    print("  ESP32 Real-Time Collector")
    print("=" * 62)
    print(f"  Attack CSV   : {ATTACK_CSV_PATH}")
    print(f"  Normal CSV   : {NORMAL_CSV_PATH} (deduplicated)")
    print(f"  Sensor CSV   : {SENSOR_CSV_PATH}")
    print(f"  MQTT         : {args.broker}:{args.port}")
    print(f"  ESP32 timeout: {ESP32_TIMEOUT_SECS}s  grace: {ESP32_GRACE_SECS}s")
    print("=" * 62 + "\n")

    _ensure_csv_headers()
    if not REALTIME_XLS_PATH.exists():
        _create_excel()
    if not SENSOR_XLS_PATH.exists():
        _create_sensor_excel()

    threading.Thread(target=_flush_loop,          daemon=True).start()
    threading.Thread(target=_stats_loop,          daemon=True).start()
    threading.Thread(target=_esp32_watchdog_loop, daemon=True).start()

    threads = []
    if not args.no_mqtt:
        t = threading.Thread(target=start_mqtt_listener,
                             args=(args.broker, args.port), daemon=True)
        t.start(); threads.append(t)
    if not args.no_http:
        t = threading.Thread(target=start_http_listener,
                             args=(args.http_host, args.http_port), daemon=True)
        t.start(); threads.append(t)
    if not threads:
        log.error("Nothing to run — use --no-mqtt or --no-http to disable listeners")
        sys.exit(1)

    log.info("Collector running. Press Ctrl+C to stop.\n")
    while _running:
        time.sleep(1)


if __name__ == "__main__":
    main()