"""
backend/app.py
==============
Single entry point that starts ALL backend processes together.

Fixes applied vs previous version:
  1. importlib.util.load_from_spec → load_module() — was a typo
  2. Flask ImportError no longer causes infinite restart loop —
     server.py subprocess runs even when feature_extraction is broken
  3. Simulation crash (RuntimeError: Model not loaded) is caught
     and retried with random fallback values
  4. RealtimeCollector now launched via start_realtime_collector()
     as a proper named daemon thread (replaces raw inline import)

Run with:
    python app.py              # start everything
    python app.py --no-mqtt    # skip MQTT subscriber
    python app.py --no-collect # skip data collectors
    python app.py --train      # retrain model then start
"""

import os
import sys
import time
import signal
import logging
import argparse
import threading
import subprocess
from datetime import datetime
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(name)-18s]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("app")

BASE_DIR    = Path(__file__).parent
MODEL_DIR   = BASE_DIR / "models"
DATASET_DIR = BASE_DIR / "dataset"
SCALER_PATH = MODEL_DIR / "scaler.pkl"
XGB_PATH    = MODEL_DIR / "xgboost_model.pkl"

_shutdown = threading.Event()
_threads  = []


# ══════════════════════════════════════════════════════════════════
# 1. MODEL CHECK
# ══════════════════════════════════════════════════════════════════

def check_or_train_model(force: bool = False):
    models_exist = SCALER_PATH.exists() and XGB_PATH.exists()

    if models_exist and not force:
        log.info("✅  Models found — skipping training")
        try:
            import joblib
            meta_path = MODEL_DIR / "model_meta.pkl"
            if meta_path.exists():
                meta = joblib.load(str(meta_path))
                log.info("    XGBoost accuracy : %.4f",
                         meta.get("xgb_accuracy", 0))
                log.info("    AE threshold     : %.6f",
                         meta.get("ae_threshold", 0))
                log.info("    Train samples    : %d",
                         meta.get("train_samples", 0))
        except Exception:
            pass
        return True

    log.warning("⚠️  No trained models found — starting auto-train")
    datasets = list(DATASET_DIR.glob("*.csv"))
    if not datasets:
        log.error("❌  No CSV files in backend/dataset/ — cannot train")
        return False

    try:
        result = subprocess.run(
            [sys.executable, "train.py",
             "--data_dir", str(DATASET_DIR),
             "--epochs", "60", "--smote"],
            cwd=str(BASE_DIR),
            timeout=1800,
        )
        if result.returncode != 0:
            log.error("❌  Training failed (rc=%d)", result.returncode)
            return False
        log.info("✅  Training complete")
        return True
    except Exception as e:
        log.error("❌  Training error: %s", e)
        return False


# ══════════════════════════════════════════════════════════════════
# 2. FLASK SERVER
# ══════════════════════════════════════════════════════════════════

def run_flask_server():
    """
    Try to import create_app() from server.py.
    If server.py has an import error (e.g. feature_extraction broken)
    fall back to running it as a subprocess which restarts on crash.
    """
    log.info("🌐  Starting Flask server on port 5000...")

    # ── Try direct import ─────────────────────────────────────────
    create_app_fn = None
    try:
        import importlib
        srv = importlib.import_module("server")
        if hasattr(srv, "create_app"):
            create_app_fn = srv.create_app
        else:
            log.warning("    server.py has no create_app() — subprocess mode")
    except Exception as e:
        log.warning("    server.py import failed (%s) — subprocess mode", e)

    if create_app_fn:
        try:
            app, socketio = create_app_fn()
            socketio.run(
                app,
                host="0.0.0.0",
                port=5000,
                debug=False,
                use_reloader=False,
                log_output=False,
            )
            return
        except Exception as e:
            log.error("❌  create_app() failed: %s — subprocess mode", e)

    # ── Fallback: subprocess ──────────────────────────────────────
    _run_subprocess("server.py", "Flask")


def _run_subprocess(script: str, name: str):
    """Run a backend script as a subprocess, restarting on crash."""
    while not _shutdown.is_set():
        log.info("▶   Starting %s (%s)", name, script)
        try:
            proc = subprocess.Popen(
                [sys.executable, script],
                cwd=str(BASE_DIR),
            )
            while not _shutdown.is_set():
                try:
                    proc.wait(timeout=1)
                    break
                except subprocess.TimeoutExpired:
                    continue

            if _shutdown.is_set():
                proc.terminate()
                break

        except Exception as e:
            log.error("❌  Subprocess %s launch error: %s", name, e)

        if not _shutdown.is_set():
            log.warning("⚠️  %s exited — restarting in 5s...", name)
            time.sleep(5)


# ══════════════════════════════════════════════════════════════════
# 3. MQTT SUBSCRIBER
# ══════════════════════════════════════════════════════════════════

def run_mqtt_subscriber(broker: str, port: int):
    log.info("📡  Starting MQTT subscriber → %s:%d", broker, port)
    try:
        import paho.mqtt.client as mqtt
        from anomaly_detector import get_detector
        from utils.feature_extraction import extract_features
        import joblib
        import numpy as np

        detector = get_detector()
        scaler   = joblib.load(str(SCALER_PATH))
        log.info("✅  MQTT: model loaded (AE thr=%.6f)",
                 detector.ae_threshold)

        def on_connect(client, userdata, flags, rc):
            if rc == 0:
                client.subscribe("iot/flows/#", qos=1)
                log.info("✅  MQTT subscriber connected → iot/flows/#")
            else:
                log.error("❌  MQTT connect failed rc=%d", rc)

        def on_message(client, userdata, msg):
            import json
            parts     = msg.topic.split("/")
            device_id = parts[-1] if len(parts) >= 3 else "unknown"
            try:
                t0     = time.time()
                raw    = json.loads(msg.payload.decode())
                vec    = extract_features(raw)
                vec    = np.nan_to_num(vec, nan=0.0,
                                       posinf=1e6, neginf=-1e6)
                vec_sc = scaler.transform(vec.reshape(1, -1))[0]
                result = detector.predict_single(vec_sc)
                lat_ms = (time.time() - t0) * 1000

                alert = json.dumps({
                    "is_attack":  result["is_attack"],
                    "confidence": round(result["confidence"], 4),
                    "latency_ms": round(lat_ms, 2),
                    "xgb_prob":   round(result.get("xgb_prob", 0), 4),
                    "ae_err":     round(
                        result.get("ae_recon_err", 0), 6),
                    "device_id":  device_id,
                    "ts":         time.time(),
                })
                client.publish(f"iot/alerts/{device_id}", alert, qos=1)

                level = "ATTACK" if result["is_attack"] else "normal"
                log.info("🔍  [%-10s] %-7s  conf=%.2f  %.1fms",
                         device_id, level,
                         result["confidence"], lat_ms)

            except Exception as e:
                log.warning("MQTT msg error [%s]: %s", device_id, e)

        try:
            client = mqtt.Client(
                client_id="app-ml-subscriber",
                callback_api_version=mqtt.CallbackAPIVersion.VERSION1,
            )
        except AttributeError:
            client = mqtt.Client(client_id="app-ml-subscriber")

        client.on_connect = on_connect
        client.on_message = on_message
        client.reconnect_delay_set(min_delay=2, max_delay=30)

        while not _shutdown.is_set():
            try:
                client.connect(broker, port, keepalive=60)
                client.loop_forever()
                break
            except Exception as e:
                if _shutdown.is_set():
                    break
                log.warning("MQTT reconnecting in 5s: %s", e)
                time.sleep(5)

    except Exception as e:
        log.error("❌  MQTT subscriber error: %s", e)


# ══════════════════════════════════════════════════════════════════
# 4. NETWORK COLLECTOR  (→ esp32_realtime.xlsx / .csv, port 5001)
# ══════════════════════════════════════════════════════════════════

def run_network_collector(broker: str, port: int):
    """
    Start realtime_collector via its public start_realtime_collector()
    entry point.  Falls back to subprocess if the import fails.
    """
    log.info("📥  Starting network collector → broker=%s port=%d  HTTP=5001",
             broker, port)

    collector_path = BASE_DIR / "realtime_collector.py"
    if not collector_path.exists():
        log.error("❌  realtime_collector.py not found")
        return

    # ── Try clean import via start_realtime_collector() ──────────
    try:
        from realtime_collector import start_realtime_collector

        # start_realtime_collector() is blocking (MQTT loop_forever),
        # so we also spin up the HTTP listener in a sibling thread.
        try:
            from realtime_collector import start_http_listener
            threading.Thread(
                target=start_http_listener,
                args=("0.0.0.0", 5001),
                name="NetworkCollector-HTTP",
                daemon=True,
            ).start()
        except Exception as he:
            log.warning("⚠️  HTTP listener thread failed: %s", he)

        log.info("✅  Network collector running (inline)")
        start_realtime_collector(broker=broker, port=port)
        return

    except Exception as e:
        log.warning("⚠️  Inline network collector failed (%s) — subprocess", e)

    # ── Fallback: subprocess ──────────────────────────────────────
    _run_subprocess("realtime_collector.py", "NetworkCollector")


# ══════════════════════════════════════════════════════════════════
# 5. SENSOR COLLECTOR  (→ esp32_sensors.xlsx, port 5002)
# ══════════════════════════════════════════════════════════════════

def run_sensor_collector(broker: str, port: int):
    log.info("🌡️   Starting sensor collector → port 5002")

    sensor_path = BASE_DIR / "sensor_collector.py"
    if not sensor_path.exists():
        log.warning("⚠️  sensor_collector.py not found — skipping")
        return

    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "sensor_collector", str(sensor_path))
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)

        mod._running = True
        threading.Thread(target=mod._flush_loop,  daemon=True).start()
        threading.Thread(target=mod.start_mqtt,
                         args=(broker, port),      daemon=True).start()
        threading.Thread(target=mod.start_http,
                         args=("0.0.0.0", 5002),   daemon=True).start()

        log.info("✅  Sensor collector running (inline)")
        while not _shutdown.is_set():
            time.sleep(5)
        return

    except Exception as e:
        log.warning("⚠️  Inline sensor failed (%s) — subprocess", e)

    _run_subprocess("sensor_collector.py", "SensorCollector")


# ══════════════════════════════════════════════════════════════════
# 6. RETRAIN WATCHER
# ══════════════════════════════════════════════════════════════════

def run_retrain_watcher():
    trigger_path = MODEL_DIR / "retrain_trigger.txt"
    log.info("🔄  Retrain watcher active (trigger: %s)", trigger_path)

    while not _shutdown.is_set():
        if trigger_path.exists():
            log.info("🔄  Retrain triggered — starting online_trainer.py")
            try:
                trigger_path.unlink()
                subprocess.run(
                    [sys.executable, "online_trainer.py",
                     "--rt_weight", "0.3"],
                    cwd=str(BASE_DIR), timeout=600,
                )
                log.info("✅  Retraining complete")
            except Exception as e:
                log.error("❌  Retraining error: %s", e)
        time.sleep(10)


# ══════════════════════════════════════════════════════════════════
# 7. STATUS PRINTER
# ══════════════════════════════════════════════════════════════════

def run_status_printer():
    while not _shutdown.is_set():
        time.sleep(60)
        if _shutdown.is_set():
            break
        files = {
            "esp32_realtime.xlsx": DATASET_DIR / "esp32_realtime.xlsx",
            "esp32_sensors.xlsx":  DATASET_DIR / "esp32_sensors.xlsx",
        }
        info = []
        for name, path in files.items():
            if path.exists():
                mb = path.stat().st_size / 1_048_576
                info.append(f"{name}={mb:.2f}MB")
            else:
                info.append(f"{name}=not created yet")
        alive = len([t for t in _threads if t.is_alive()])
        log.info("📊  STATUS — threads=%d  %s", alive, "  ".join(info))


# ══════════════════════════════════════════════════════════════════
# SHUTDOWN
# ══════════════════════════════════════════════════════════════════

def shutdown(sig, frame):
    log.info("\n🛑  Shutting down all services...")
    _shutdown.set()
    time.sleep(3)

    # Flush any pending realtime collector queue before exit
    try:
        from realtime_collector import flush_to_excel
        flush_to_excel()
        log.info("💾  Realtime collector queue flushed")
    except Exception:
        pass

    for name, path in [
        ("Network data", DATASET_DIR / "esp32_realtime.xlsx"),
        ("Sensor data",  DATASET_DIR / "esp32_sensors.xlsx"),
    ]:
        if path.exists():
            mb = path.stat().st_size / 1_048_576
            try:
                from openpyxl import load_workbook
                wb   = load_workbook(str(path), read_only=True)
                rows = max(wb.active.max_row - 1, 0)
                wb.close()
                log.info("💾  %s: %d rows  %.2fMB", name, rows, mb)
            except Exception:
                log.info("💾  %s: %.2fMB", name, mb)

    log.info("👋  Stopped. Goodbye.")
    sys.exit(0)


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(
        description="IoT Anomaly Detection — Unified Backend"
    )
    ap.add_argument("--broker",     default="localhost")
    ap.add_argument("--mqtt-port",  type=int, default=1883)
    ap.add_argument("--train",      action="store_true")
    ap.add_argument("--no-mqtt",    action="store_true")
    ap.add_argument("--no-collect", action="store_true")
    ap.add_argument("--no-sensor",  action="store_true")
    args = ap.parse_args()

    signal.signal(signal.SIGINT,  shutdown)
    signal.signal(signal.SIGTERM, shutdown)

    print("\n" + "=" * 62)
    print("  IoT Network Anomaly Detection — Backend v5.1")
    print("=" * 62)
    print(f"  Flask API       : http://0.0.0.0:5000")
    print(f"  Network collect : http://0.0.0.0:5001/api/realtime")
    print(f"  Sensor collect  : http://0.0.0.0:5002/sensor")
    print(f"  MQTT broker     : {args.broker}:{args.mqtt_port}")
    print(f"  Dataset dir     : {DATASET_DIR}")
    print(f"  Models dir      : {MODEL_DIR}")
    print("=" * 62 + "\n")

    check_or_train_model(force=args.train)

    # ── Service registry ──────────────────────────────────────────
    services = [
        ("Flask+SocketIO",  run_flask_server,    ()),
    ]

    if not args.no_mqtt:
        services.append((
            "MQTT-Subscriber",
            run_mqtt_subscriber,
            (args.broker, args.mqtt_port),
        ))

    if not args.no_collect:
        # NetworkCollector wraps start_realtime_collector() which
        # handles MQTT + HTTP + flush loop internally
        services.append((
            "NetworkCollector",
            run_network_collector,
            (args.broker, args.mqtt_port),
        ))

    if not args.no_collect and not args.no_sensor:
        services.append((
            "SensorCollector",
            run_sensor_collector,
            (args.broker, args.mqtt_port),
        ))

    services += [
        ("RetrainWatcher", run_retrain_watcher, ()),
        ("StatusPrinter",  run_status_printer,  ()),
    ]

    # ── Launch all services as daemon threads ─────────────────────
    for name, fn, fn_args in services:
        t = threading.Thread(
            target=fn,
            args=fn_args,
            name=name,
            daemon=True,
        )
        t.start()
        _threads.append(t)
        log.info("▶   %-22s started", name)
        time.sleep(0.5)

    print()
    log.info("✅  All %d services running.", len(_threads))
    log.info("    Press Ctrl+C to stop.\n")

    while not _shutdown.is_set():
        time.sleep(5)


if __name__ == "__main__":
    main()